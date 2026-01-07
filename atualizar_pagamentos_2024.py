"""
Script para atualizar pagamentos de 2024 a partir da aba 2024 do arquivo 2025.xlsx
Marca como PAGO os honorários onde há data e forma de pagamento preenchidos.
"""
import openpyxl
import os
import sys
import re

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), 'honorarios'))
from database import get_connection

# Mapeamento de colunas para meses
COLUNAS_MESES = {
    3: 1,   # C = Janeiro
    4: 2,   # D = Fevereiro
    5: 3,   # E = Março
    6: 4,   # F = Abril
    7: 5,   # G = Maio
    8: 6,   # H = Junho
    9: 7,   # I = Julho
    10: 8,  # J = Agosto
    11: 9,  # K = Setembro
    12: 10, # L = Outubro
    13: 11, # M = Novembro
    14: 13, # N = 13º Salário
    15: 12, # O = Dezembro
}

MESES_ABREV = {
    'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4, 'mai': 5, 'jun': 6,
    'jul': 7, 'ago': 8, 'set': 9, 'out': 10, 'nov': 11, 'dez': 12,
}

def parse_data(data_str, ano):
    """Converte data para formato YYYY-MM-DD"""
    if not data_str:
        return None
    try:
        data_str = str(data_str).lower().strip()
        
        # Formato: DD/mes (ex: 09/dez)
        match = re.match(r'(\d{1,2})[/\\-](\w+)', data_str)
        if match:
            dia = int(match.group(1))
            mes_str = match.group(2)
            mes = MESES_ABREV.get(mes_str[:3], None)
            if mes:
                return f"{ano}-{mes:02d}-{dia:02d}"
        
        # Formato: DD/MM/YYYY ou DD/MM
        match = re.match(r'(\d{1,2})[/\\-](\d{1,2})(?:[/\\-](\d{2,4}))?', data_str)
        if match:
            dia = int(match.group(1))
            mes = int(match.group(2))
            ano_data = int(match.group(3)) if match.group(3) else ano
            if ano_data < 100:
                ano_data += 2000
            return f"{ano_data}-{mes:02d}-{dia:02d}"
        
        return None
    except:
        return None

def atualizar_pagamentos(arquivo, aba, ano):
    """Atualiza pagamentos do ano specificado a partir da aba"""
    print(f"\n📂 Processando: {arquivo} - Aba: {aba} - Ano: {ano}")
    
    if not os.path.exists(arquivo):
        print(f"[ERRO] Arquivo não encontrado: {arquivo}")
        return 0, 0
    
    wb = openpyxl.load_workbook(arquivo, data_only=True)
    
    if aba not in wb.sheetnames:
        print(f"[ERRO] Aba '{aba}' não encontrada. Abas disponíveis: {wb.sheetnames}")
        return 0, 0
    
    ws = wb[aba]
    
    conn = get_connection()
    cursor = conn.cursor()
    
    # Carregar mapeamento de clientes por código
    cursor.execute("SELECT id, codigo_interno, UPPER(nome) as nome FROM clientes")
    clientes_db = {}
    clientes_nome = {}
    for c in cursor.fetchall():
        if c['codigo_interno']:
            clientes_db[str(c['codigo_interno']).strip()] = c['id']
        clientes_nome[c['nome']] = c['id']
    
    atualizados = 0
    erros = 0
    nao_encontrados = 0
    
    linha = 2  # Começar na linha 2 (pular cabeçalho)
    
    while linha <= ws.max_row:
        codigo_cliente = ws.cell(row=linha, column=1).value
        nome_cliente = ws.cell(row=linha, column=2).value
        
        if not codigo_cliente or not nome_cliente:
            linha += 4
            continue
        
        try:
            codigo_int = int(codigo_cliente)
        except (ValueError, TypeError):
            linha += 4
            continue
        
        # Encontrar cliente
        codigo_str = str(codigo_cliente).strip()
        nome_str = str(nome_cliente).strip() if nome_cliente else ""
        
        cliente_id = None
        if codigo_str in clientes_db:
            cliente_id = clientes_db[codigo_str]
        elif nome_str.upper() in clientes_nome:
            cliente_id = clientes_nome[nome_str.upper()]
        
        if not cliente_id:
            # Criar cliente automaticamente como inativo
            try:
                cursor.execute("""
                    INSERT INTO clientes (nome, codigo_interno, email, ativo)
                    VALUES (?, ?, '', 0)
                """, (nome_str, codigo_str))
                cliente_id = cursor.lastrowid
                clientes_db[codigo_str] = cliente_id
                clientes_nome[nome_str.upper()] = cliente_id
                print(f"  [NOVO] Cliente criado: {codigo_str} - {nome_str}")
            except Exception as e:
                print(f"  [ERRO] Falha ao criar cliente: {e}")
                nao_encontrados += 1
                linha += 4
                continue
        
        # Processar cada mês
        for col, mes in COLUNAS_MESES.items():
            # Valor (mesma linha do código/nome)
            valor_cell = ws.cell(row=linha, column=col).value
            
            # Data (linha + 2)
            data_cell = ws.cell(row=linha + 2, column=col).value
            data_pagamento = parse_data(data_cell, ano)
            
            # Forma de pagamento (linha + 3)
            forma_cell = ws.cell(row=linha + 3, column=col).value
            forma_pagamento = str(forma_cell).strip() if forma_cell else None
            
            # Parsear valor
            valor = None
            if valor_cell:
                if isinstance(valor_cell, (int, float)):
                    valor = float(valor_cell)
                else:
                    valor_str = str(valor_cell).replace('R$', '').replace(' ', '').replace('.', '').replace(',', '.')
                    try:
                        valor = float(valor_str)
                    except:
                        valor = None
            
            # Se tem VALOR preenchido, processar como PAGO
            if valor and valor > 0:
                try:
                    # Criar honorário se não existir
                    cursor.execute("""
                        INSERT OR IGNORE INTO honorarios (cliente_id, ano, mes, valor, status)
                        VALUES (?, ?, ?, ?, 'PENDENTE')
                    """, (cliente_id, ano, mes, valor))
                    
                    # Atualizar para PAGO
                    cursor.execute("""
                        UPDATE honorarios 
                        SET status = 'PAGO', valor = ?, data_pagamento = ?, forma_pagamento = ?
                        WHERE cliente_id = ? AND ano = ? AND mes = ?
                    """, (valor, data_pagamento, forma_pagamento, cliente_id, ano, mes))
                    
                    if cursor.rowcount > 0:
                        atualizados += 1
                except Exception as e:
                    print(f"  [ERRO] Cliente {codigo_str}, Mês {mes}: {e}")
                    erros += 1
        
        linha += 4
    
    conn.commit()
    conn.close()
    
    print(f"  ✅ Atualizados: {atualizados}")
    print(f"  ⚠️ Clientes não encontrados: {nao_encontrados}")
    print(f"  ❌ Erros: {erros}")
    
    return atualizados, erros

if __name__ == "__main__":
    print("\n" + "="*60)
    print("   ATUALIZAÇÃO DE PAGAMENTOS - ANO 2024")
    print("="*60)
    
    atualizados, erros = atualizar_pagamentos("2025.xlsx", "2024", 2024)
    
    print("\n" + "="*60)
    print(f"   CONCLUÍDO!")
    print(f"   Total atualizados para PAGO: {atualizados}")
    print(f"   Total erros: {erros}")
    print("="*60)
