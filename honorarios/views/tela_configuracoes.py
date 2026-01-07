"""
Tela de Configurações - Visual Moderno
Suporte light/dark
"""
import flet as ft
from datetime import datetime
import os

from database import listar_clientes, adicionar_cliente, atualizar_cliente, get_config, set_config, get_connection, set_valor_honorario_ano, listar_valores_honorarios_cliente, criar_honorarios_ano_cliente
from utils.theme import CORES
from utils.toast import toast_success, toast_error, toast_warning


def criar_tela_configuracoes(page: ft.Page, usuario_logado, theme, VERSION="0.1"):
    """Tela de configurações moderna"""
    
    # Cores dinâmicas
    BG = theme.get_bg()
    SURFACE = theme.get_surface()
    SURFACE_HOVER = theme.get_surface_hover()
    INPUT_BG = theme.get_input_bg()
    INPUT_BORDER = theme.get_input_border()
    TEXT_PRIMARY = theme.get_text_color()
    TEXT_SECONDARY = theme.get_text_secondary()
    ACCENT = CORES["accent"]
    SUCCESS = CORES["success"]
    WARNING = CORES["warning"]
    DANGER = CORES["danger"]
    
    clientes_list = ft.Column([], spacing=4, scroll=ft.ScrollMode.AUTO)
    
    def atualizar_clientes(filtro=""):
        clientes_list.controls.clear()
        clientes = sorted(listar_clientes(), key=lambda c: int(c['codigo_interno']) if c['codigo_interno'] and c['codigo_interno'].isdigit() else float('inf'))
        
        for cliente in clientes:
            if filtro and filtro.lower() not in cliente['nome'].lower():
                continue
            
            codigo = cliente['codigo_interno'] or str(cliente['id'])
            # Tentar acessar valor_honorario de forma segura
            try:
                valor_hon = cliente['valor_honorario']
            except (KeyError, IndexError):
                valor_hon = None
            valor_str = f"R$ {valor_hon:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") if valor_hon else "-"
            
            clientes_list.controls.append(
                ft.Container(
                    content=ft.Row([
                        ft.Text(codigo, size=10, width=50, color=TEXT_SECONDARY),
                        ft.Text(cliente['nome'][:25], size=11, color=TEXT_PRIMARY, expand=True),
                        ft.Text(cliente['cnpj'] or cliente['cpf'] or "-", size=9, color=TEXT_SECONDARY, width=120),
                        ft.Text(valor_str, size=10, weight=ft.FontWeight.BOLD, color=SUCCESS if valor_hon else TEXT_SECONDARY, width=90),
                        ft.IconButton(
                            ft.Icons.EDIT,
                            icon_color=ACCENT,
                            icon_size=16,
                            on_click=lambda e, c=cliente: abrir_modal(c),
                        ),
                    ]),
                    padding=ft.padding.symmetric(horizontal=10, vertical=6),
                    bgcolor=SURFACE,
                    border_radius=8,
                    border=ft.border.all(1, theme.get_border()),
                )
            )
        page.update()
    
    def abrir_modal(cliente=None):
        is_novo = cliente is None
        
        nome = ft.TextField(label="Nome *", width=350, value="" if is_novo else cliente['nome'],
                           bgcolor=INPUT_BG, border_color=INPUT_BORDER, color=TEXT_PRIMARY)
        codigo = ft.TextField(label="Código", width=120, value="" if is_novo else (cliente['codigo_interno'] or ""),
                             bgcolor=INPUT_BG, border_color=INPUT_BORDER, color=TEXT_PRIMARY)
        cnpj = ft.TextField(label="CNPJ", width=170, value="" if is_novo else (cliente['cnpj'] or ""),
                           bgcolor=INPUT_BG, border_color=INPUT_BORDER, color=TEXT_PRIMARY)
        cpf = ft.TextField(label="CPF", width=140, value="" if is_novo else (cliente['cpf'] or ""),
                          bgcolor=INPUT_BG, border_color=INPUT_BORDER, color=TEXT_PRIMARY)
        endereco = ft.TextField(label="Endereço", width=350, value="" if is_novo else (cliente['endereco'] or ""),
                               bgcolor=INPUT_BG, border_color=INPUT_BORDER, color=TEXT_PRIMARY)
        telefone = ft.TextField(label="Telefone", width=170, value="" if is_novo else (cliente['telefone'] or ""),
                               bgcolor=INPUT_BG, border_color=INPUT_BORDER, color=TEXT_PRIMARY)
        email = ft.TextField(label="Email", width=250, value="" if is_novo else (cliente['email'] or ""),
                            bgcolor=INPUT_BG, border_color=INPUT_BORDER, color=TEXT_PRIMARY)
        
        # Obter valor_honorario de forma segura
        valor_hon_atual = ""
        if not is_novo:
            try:
                valor_hon_atual = str(cliente['valor_honorario'] or "")
            except (KeyError, IndexError):
                valor_hon_atual = ""
        
        valor_honorario = ft.TextField(
            label="Valor Honorário (R$)", width=150, 
            value=valor_hon_atual,
            bgcolor=INPUT_BG, border_color=INPUT_BORDER, color=TEXT_PRIMARY,
            keyboard_type=ft.KeyboardType.NUMBER,
        )
        
        def salvar(e):
            if not nome.value:
                toast_warning(page, "Nome é obrigatório!")
                return
            
            try:
                # Converter valor para float
                valor_hon = None
                if valor_honorario.value:
                    try:
                        valor_hon = float(valor_honorario.value.replace(",", "."))
                    except:
                        pass
                
                if is_novo:
                    adicionar_cliente(
                        nome=nome.value, cnpj=cnpj.value or None, cpf=cpf.value or None,
                        endereco=endereco.value or None, telefone=telefone.value or None,
                        email=email.value or None, codigo_interno=codigo.value or None,
                        valor_honorario=valor_hon,
                    )
                    toast_success(page, "Cliente adicionado!")
                else:
                    atualizar_cliente(
                        cliente_id=cliente['id'], nome=nome.value, cnpj=cnpj.value or None,
                        cpf=cpf.value or None, endereco=endereco.value or None,
                        telefone=telefone.value or None, email=email.value or None,
                        codigo_interno=codigo.value or None, valor_honorario=valor_hon,
                    )
                    toast_success(page, "Cliente atualizado!")
                
                dlg.open = False
                atualizar_clientes()
            except Exception as ex:
                toast_error(page, f"Erro: {str(ex)}")
        
        def fechar(e):
            dlg.open = False
            page.update()
        
        # === ABA VALORES POR ANO (somente edição) ===
        valores_por_ano = {}
        ano_atual = datetime.now().year
        campos_anos = {}
        
        if not is_novo:
            # Carregar valores existentes
            valores_existentes = listar_valores_honorarios_cliente(cliente['id'])
            for ano_val, valor_val in valores_existentes:
                valores_por_ano[ano_val] = valor_val
        
        # Criar campos de 2021 a ano_atual + 1
        for ano in range(2021, ano_atual + 2):
            valor_existente = valores_por_ano.get(ano, "")
            campos_anos[ano] = ft.TextField(
                label=str(ano), width=100,
                value=str(valor_existente) if valor_existente else "",
                bgcolor=INPUT_BG, border_color=INPUT_BORDER, color=TEXT_PRIMARY,
                keyboard_type=ft.KeyboardType.NUMBER,
            )
        
        def gerar_honorarios_ano(e, ano_selecionado):
            """Gera 12 honorários pendentes para o ano selecionado"""
            if is_novo:
                toast_warning(page, "Salve o cliente primeiro!")
                return
            
            valor_str = campos_anos[ano_selecionado].value
            if not valor_str:
                toast_warning(page, f"Informe o valor para {ano_selecionado}")
                return
            
            try:
                valor = float(valor_str.replace(",", "."))
                criar_honorarios_ano_cliente(cliente['id'], ano_selecionado, valor)
                toast_success(page, f"Honorários de {ano_selecionado} gerados!")
            except Exception as ex:
                toast_error(page, f"Erro: {ex}")
        
        def salvar_valores_anos(e):
            """Salva todos os valores por ano"""
            if is_novo:
                toast_warning(page, "Salve o cliente primeiro!")
                return
            
            for ano, campo in campos_anos.items():
                if campo.value:
                    try:
                        valor = float(campo.value.replace(",", "."))
                        set_valor_honorario_ano(cliente['id'], ano, valor)
                    except:
                        pass
            toast_success(page, "Valores por ano salvos!")
        
        # Layout da aba Valores por Ano
        rows_anos = []
        anos_lista = list(campos_anos.keys())
        for i in range(0, len(anos_lista), 3):
            row_items = []
            for ano in anos_lista[i:i+3]:
                campo = campos_anos[ano]
                btn_gerar = ft.IconButton(
                    icon=ft.Icons.ADD_CIRCLE_OUTLINE,
                    tooltip=f"Gerar Honorários {ano}",
                    icon_size=20,
                    icon_color=ACCENT,
                    on_click=lambda e, a=ano: gerar_honorarios_ano(e, a),
                )
                row_items.append(ft.Row([campo, btn_gerar], spacing=2))
            rows_anos.append(ft.Row(row_items, spacing=15))
        
        aba_dados = ft.Tab(
            text="📋 Dados",
            content=ft.Container(
                padding=10,
                content=ft.Column([
                    nome,
                    ft.Row([codigo, cnpj], spacing=10),
                    ft.Row([cpf, telefone], spacing=10),
                    endereco,
                    ft.Row([email, valor_honorario], spacing=10),
                ], spacing=10),
            )
        )
        
        aba_valores = ft.Tab(
            text="💰 Valores/Ano",
            content=ft.Container(
                padding=10,
                content=ft.Column([
                    ft.Text("Valor do honorário para cada ano:", size=12, color=TEXT_SECONDARY),
                    *rows_anos,
                    ft.ElevatedButton(
                        "💾 Salvar Valores",
                        bgcolor=SUCCESS, color="white",
                        on_click=salvar_valores_anos,
                    ),
                ], spacing=10),
            )
        )
        
        tabs = ft.Tabs(
            selected_index=0,
            tabs=[aba_dados, aba_valores] if not is_novo else [aba_dados],
            expand=True,
        )
        
        dlg = ft.AlertDialog(
            modal=True,
            title=ft.Text("➕ Novo Cliente" if is_novo else "✏️ Editar", weight=ft.FontWeight.BOLD, color=TEXT_PRIMARY),
            bgcolor=BG,
            content=ft.Container(
                width=500,
                height=350,
                content=tabs,
            ),
            actions=[
                ft.TextButton("Cancelar", on_click=fechar),
                ft.ElevatedButton("Salvar Dados", bgcolor=SUCCESS, color=TEXT_PRIMARY, on_click=salvar),
            ],
        )
        page.overlay.append(dlg)
        dlg.open = True
        page.update()
    
    # ═══ EXPORTAR CLIENTES ═══
    wb_exportar = [None]  # Para armazenar o workbook temporariamente
    
    def preparar_exportacao(e):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Clientes"
            
            # Anos para colunas de valores
            anos = list(range(2021, datetime.now().year + 2))  # 2021 até ano_atual + 1
            
            # Cabeçalhos
            headers = ["ID", "Código", "Nome", "CNPJ", "CPF", "Endereço", "Telefone", "Email", "Valor Honorário"]
            # Adicionar colunas de anos
            for ano in anos:
                headers.append(f"Valor {ano}")
            
            header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Dados
            clientes = sorted(listar_clientes(), key=lambda c: int(c['codigo_interno']) if c['codigo_interno'] and c['codigo_interno'].isdigit() else float('inf'))
            for row, cliente in enumerate(clientes, 2):
                ws.cell(row=row, column=1, value=cliente['id'])
                ws.cell(row=row, column=2, value=cliente['codigo_interno'] or "")
                ws.cell(row=row, column=3, value=cliente['nome'])
                ws.cell(row=row, column=4, value=cliente['cnpj'] or "")
                ws.cell(row=row, column=5, value=cliente['cpf'] or "")
                ws.cell(row=row, column=6, value=cliente['endereco'] or "")
                ws.cell(row=row, column=7, value=cliente['telefone'] or "")
                ws.cell(row=row, column=8, value=cliente['email'] or "")
                try:
                    ws.cell(row=row, column=9, value=cliente['valor_honorario'] or "")
                except:
                    ws.cell(row=row, column=9, value="")
                
                # Valores por ano
                valores_cliente = listar_valores_honorarios_cliente(cliente['id'])
                valores_dict = {ano: valor for ano, valor in valores_cliente}
                
                for col_idx, ano in enumerate(anos, 10):
                    valor_ano = valores_dict.get(ano, "")
                    ws.cell(row=row, column=col_idx, value=valor_ano if valor_ano else "")
            
            # Ajustar largura
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 35
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 40
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 30
            ws.column_dimensions['I'].width = 15
            # Colunas de anos
            from openpyxl.utils import get_column_letter
            for i in range(len(anos)):
                ws.column_dimensions[get_column_letter(10 + i)].width = 12
            
            wb_exportar[0] = wb
            # Abrir seletor de pasta
            export_picker.save_file(
                file_name=f"clientes_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                allowed_extensions=["xlsx"]
            )
        except ImportError:
            toast_error(page, "Instale: pip install openpyxl")
        except Exception as ex:
            toast_error(page, f"Erro: {str(ex)}")
    
    def salvar_exportacao(e: ft.FilePickerResultEvent):
        if not e.path or not wb_exportar[0]:
            return
        try:
            wb_exportar[0].save(e.path)
            toast_success(page, f"Exportado: {os.path.basename(e.path)}")
            wb_exportar[0] = None
        except Exception as ex:
            toast_error(page, f"Erro ao salvar: {str(ex)}")
    
    export_picker = ft.FilePicker(on_result=salvar_exportacao)
    page.overlay.append(export_picker)
    
    # ═══ IMPORTAR CLIENTES ═══
    def importar_clientes(e: ft.FilePickerResultEvent):
        if not e.files:
            return
        
        try:
            from openpyxl import load_workbook
            
            filepath = e.files[0].path
            wb = load_workbook(filepath)
            ws = wb.active
            
            conn = get_connection()
            cursor = conn.cursor()
            
            atualizados = 0
            adicionados = 0
            valores_salvos = 0
            
            # Detectar colunas de anos no cabeçalho
            anos_colunas = {}  # {ano: coluna}
            for col in range(10, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header and "Valor" in str(header):
                    try:
                        ano = int(str(header).replace("Valor", "").strip())
                        anos_colunas[ano] = col
                    except:
                        pass
            
            for row in range(2, ws.max_row + 1):
                cliente_id = ws.cell(row=row, column=1).value
                codigo = ws.cell(row=row, column=2).value or None
                nome = ws.cell(row=row, column=3).value
                
                if not nome:  # Pular linhas sem nome
                    continue
                
                cnpj = ws.cell(row=row, column=4).value or None
                cpf = ws.cell(row=row, column=5).value or None
                endereco = ws.cell(row=row, column=6).value or None
                telefone = ws.cell(row=row, column=7).value or None
                email = ws.cell(row=row, column=8).value or None
                valor_hon = ws.cell(row=row, column=9).value
                
                if valor_hon and isinstance(valor_hon, str):
                    try:
                        valor_hon = float(valor_hon.replace(",", ".").replace("R$", "").strip())
                    except:
                        valor_hon = None
                
                if cliente_id:
                    # Atualizar cliente existente
                    cursor.execute("""
                        UPDATE clientes 
                        SET codigo_interno=?, nome=?, cnpj=?, cpf=?, endereco=?, telefone=?, email=?, valor_honorario=?
                        WHERE id=?
                    """, (codigo, nome, cnpj, cpf, endereco, telefone, email, valor_hon, cliente_id))
                    atualizados += 1
                    atual_cliente_id = cliente_id
                else:
                    # Adicionar novo cliente
                    cursor.execute("""
                        INSERT INTO clientes (codigo_interno, nome, cnpj, cpf, endereco, telefone, email, valor_honorario)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """, (codigo, nome, cnpj, cpf, endereco, telefone, email, valor_hon))
                    adicionados += 1
                    atual_cliente_id = cursor.lastrowid
                
                # Importar valores por ano
                for ano, col in anos_colunas.items():
                    valor_ano = ws.cell(row=row, column=col).value
                    if valor_ano:
                        if isinstance(valor_ano, str):
                            try:
                                valor_ano = float(valor_ano.replace(",", ".").replace("R$", "").strip())
                            except:
                                continue
                        if valor_ano and valor_ano > 0:
                            cursor.execute("""
                                INSERT OR REPLACE INTO valores_honorarios (cliente_id, ano, valor)
                                VALUES (?, ?, ?)
                            """, (atual_cliente_id, ano, valor_ano))
                            valores_salvos += 1
            
            conn.commit()
            conn.close()
            
            atualizar_clientes()
            msg = []
            if adicionados > 0:
                msg.append(f"{adicionados} adicionado(s)")
            if atualizados > 0:
                msg.append(f"{atualizados} atualizado(s)")
            if valores_salvos > 0:
                msg.append(f"{valores_salvos} valores/ano")
            toast_success(page, " | ".join(msg) if msg else "Nenhuma alteração")
        except ImportError:
            toast_error(page, "Instale: pip install openpyxl")
        except Exception as ex:
            toast_error(page, f"Erro: {str(ex)}")
    
    file_picker = ft.FilePicker(on_result=importar_clientes)
    page.overlay.append(file_picker)
    
    # Tab Empresa
    empresa_nome = ft.TextField(label="Nome da Empresa", width=400, value=get_config('empresa_nome', ''),
                               bgcolor=INPUT_BG, border_color=INPUT_BORDER, color=TEXT_PRIMARY)
    empresa_cnpj = ft.TextField(label="CNPJ", width=200, value=get_config('empresa_cnpj', ''),
                               bgcolor=INPUT_BG, border_color=INPUT_BORDER, color=TEXT_PRIMARY)
    empresa_endereco = ft.TextField(label="Endereço", width=400, value=get_config('empresa_endereco', ''),
                                   bgcolor=INPUT_BG, border_color=INPUT_BORDER, color=TEXT_PRIMARY)
    empresa_cidade = ft.TextField(label="Cidade/UF", width=250, value=get_config('empresa_cidade', ''),
                                 bgcolor=INPUT_BG, border_color=INPUT_BORDER, color=TEXT_PRIMARY)
    empresa_telefone = ft.TextField(label="Telefone", width=200, value=get_config('empresa_telefone', ''),
                                   bgcolor=INPUT_BG, border_color=INPUT_BORDER, color=TEXT_PRIMARY)
    empresa_email = ft.TextField(label="Email", width=300, value=get_config('empresa_email', ''),
                                bgcolor=INPUT_BG, border_color=INPUT_BORDER, color=TEXT_PRIMARY)
    email_senha_app = ft.TextField(label="Senha App Gmail", width=200, value=get_config('email_senha_app', ''),
                                  bgcolor=INPUT_BG, border_color=INPUT_BORDER, color=TEXT_PRIMARY,
                                  password=True, can_reveal_password=True)
    
    # Template de email padrão
    template_padrao = """Prezado(a) {cliente_nome},

Segue em anexo o recibo de honorários contábeis referente a {mes}/{ano}.

Valor: {valor}

Em caso de dúvidas, entre em contato conosco.

Atenciosamente,
{empresa_nome}"""
    
    email_template = ft.TextField(
        label="Template do Email (use: {cliente_nome}, {mes}, {ano}, {valor}, {empresa_nome})",
        width=550, value=get_config('email_template', template_padrao),
        bgcolor=INPUT_BG, border_color=INPUT_BORDER, color=TEXT_PRIMARY,
        multiline=True, min_lines=6, max_lines=10,
    )
    empresa_pix = ft.TextField(label="Chave PIX", width=350, value=get_config('empresa_pix', ''),
                              bgcolor=INPUT_BG, border_color=INPUT_BORDER, color=TEXT_PRIMARY)
    
    def salvar_empresa(e):
        set_config('empresa_nome', empresa_nome.value)
        set_config('empresa_cnpj', empresa_cnpj.value)
        set_config('empresa_endereco', empresa_endereco.value)
        set_config('empresa_cidade', empresa_cidade.value)
        set_config('empresa_telefone', empresa_telefone.value)
        set_config('empresa_email', empresa_email.value)
        set_config('email_senha_app', email_senha_app.value)
        set_config('email_template', email_template.value)
        set_config('empresa_pix', empresa_pix.value)
        toast_success(page, "Dados salvos!")
    
    tabs = ft.Tabs(
        selected_index=0,
        animation_duration=150,
        indicator_color=ACCENT,
        label_color=TEXT_PRIMARY,
        unselected_label_color=TEXT_SECONDARY,
        tabs=[
            ft.Tab(
                text="👥 Clientes",
                content=ft.Container(
                    content=ft.Column([
                        ft.Row([
                            ft.TextField(
                                label="Buscar",
                                width=250,
                                prefix_icon=ft.Icons.SEARCH,
                                bgcolor=INPUT_BG,
                                border_color=INPUT_BORDER,
                                color=TEXT_PRIMARY,
                                on_change=lambda e: atualizar_clientes(e.control.value),
                            ),
                            ft.Container(expand=True),
                            ft.ElevatedButton("📥 Exportar", bgcolor=ACCENT, color=TEXT_PRIMARY, on_click=preparar_exportacao, tooltip="Exportar para Excel"),
                            ft.ElevatedButton("📤 Importar", bgcolor=WARNING, color="#1E293B", on_click=lambda e: file_picker.pick_files(allowed_extensions=["xlsx"]), tooltip="Importar Excel"),
                            ft.ElevatedButton("➕ Novo", bgcolor=SUCCESS, color=TEXT_PRIMARY, on_click=lambda e: abrir_modal()),
                        ], spacing=8),
                        ft.Container(height=10),
                        ft.Container(content=clientes_list, height=360, bgcolor=BG, border_radius=10, padding=10),
                    ]),
                    padding=20,
                ),
            ),
            ft.Tab(
                text="🏢 Empresa",
                content=ft.Container(
                    content=ft.Column([
                        ft.Text("Dados para Recibos", size=14, weight=ft.FontWeight.BOLD, color=TEXT_PRIMARY),
                        ft.Container(height=15),
                        empresa_nome,
                        ft.Row([empresa_cnpj, empresa_telefone], spacing=15),
                        empresa_endereco,
                        ft.Row([empresa_cidade, empresa_email], spacing=15),
                        ft.Row([email_senha_app, empresa_pix], spacing=15),
                        email_template,
                        ft.Container(height=20),
                        ft.ElevatedButton("💾 Salvar", bgcolor=ACCENT, color=TEXT_PRIMARY, on_click=salvar_empresa),
                    ], spacing=10),
                    padding=20,
                ),
            ),
            ft.Tab(
                text="⚙️ Sistema",
                content=ft.Container(
                    content=ft.Column([
                        ft.Text("Informações", size=14, weight=ft.FontWeight.BOLD, color=TEXT_PRIMARY),
                        ft.Container(height=15),
                        ft.Text(f"Usuário: {usuario_logado[0].get('username', 'N/A') if usuario_logado[0] else 'N/A'}", 
                               size=12, color=TEXT_SECONDARY),
                        ft.Text(f"Versão: {VERSION}", size=12, color=TEXT_SECONDARY),
                        ft.Container(height=10),
                        ft.Text("💡 Use o botão no menu lateral para alternar o tema.", size=11, color=TEXT_SECONDARY, italic=True),
                    ]),
                    padding=20,
                ),
            ),
        ],
        expand=True,
    )
    
    atualizar_clientes()
    
    return ft.Container(
        content=ft.Column([
            ft.Text("⚙️ Configurações", size=24, weight=ft.FontWeight.BOLD, color=TEXT_PRIMARY),
            ft.Container(height=10),
            tabs,
        ]),
        expand=True,
        padding=30,
        bgcolor=BG,
    )
