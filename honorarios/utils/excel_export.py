"""
Exportador de Excel Avançado - Sistema de Honorários
Relatórios formatados em Excel com filtros
"""
import os
from datetime import datetime


def exportar_excel_completo(ano, pasta_destino=None):
    """Exporta relatório completo do ano para Excel"""
    try:
        # Tentar usar openpyxl para Excel formatado
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        from database import listar_honorarios, listar_clientes, get_resumo_ano
        
        wb = Workbook()
        
        # ===== ABA 1: RESUMO =====
        ws_resumo = wb.active
        ws_resumo.title = "Resumo"
        
        resumo = get_resumo_ano(ano)
        
        # Estilos
        titulo_font = Font(bold=True, size=14, color="FFFFFF")
        titulo_fill = PatternFill(start_color="1A2957", end_color="1A2957", fill_type="solid")
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="C4A962", end_color="C4A962", fill_type="solid")
        
        # Título
        ws_resumo.merge_cells('A1:D1')
        ws_resumo['A1'] = f"Relatório de Honorários - {ano}"
        ws_resumo['A1'].font = titulo_font
        ws_resumo['A1'].fill = titulo_fill
        ws_resumo['A1'].alignment = Alignment(horizontal='center')
        
        # Métricas
        metricas = [
            ("Total de Honorários", resumo['total'] if resumo else 0),
            ("Pagos", resumo['pagos'] if resumo else 0),
            ("Pendentes", resumo['pendentes'] if resumo else 0),
            ("Valor Total", f"R$ {resumo['valor_total']:,.2f}" if resumo and resumo['valor_total'] else "R$ 0,00"),
            ("Valor Recebido", f"R$ {resumo['valor_recebido']:,.2f}" if resumo and resumo['valor_recebido'] else "R$ 0,00"),
            ("Valor Pendente", f"R$ {resumo['valor_pendente']:,.2f}" if resumo and resumo['valor_pendente'] else "R$ 0,00"),
        ]
        
        for i, (label, valor) in enumerate(metricas, start=3):
            ws_resumo[f'A{i}'] = label
            ws_resumo[f'A{i}'].font = Font(bold=True)
            ws_resumo[f'B{i}'] = valor
        
        # Ajustar larguras
        ws_resumo.column_dimensions['A'].width = 25
        ws_resumo.column_dimensions['B'].width = 20
        
        # ===== ABA 2: HONORÁRIOS =====
        ws_honorarios = wb.create_sheet("Honorários")
        
        headers = ["Cliente", "Mês", "Valor", "Status", "Observação"]
        for col, header in enumerate(headers, start=1):
            cell = ws_honorarios.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        MESES = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez", "13º"]
        honorarios = listar_honorarios(ano=ano)
        
        for i, h in enumerate(honorarios, start=2):
            ws_honorarios.cell(row=i, column=1, value=h['cliente_nome'])
            ws_honorarios.cell(row=i, column=2, value=MESES[h['mes']-1] if h['mes'] <= 13 else str(h['mes']))
            ws_honorarios.cell(row=i, column=3, value=h['valor'])
            ws_honorarios.cell(row=i, column=4, value=h['status'])
            ws_honorarios.cell(row=i, column=5, value=h['observacao'] or '')
            
            # Formatação de status
            if h['status'] == 'PAGO':
                ws_honorarios.cell(row=i, column=4).fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
            elif h['status'] == 'PENDENTE':
                ws_honorarios.cell(row=i, column=4).fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
        
        # Ajustar larguras
        ws_honorarios.column_dimensions['A'].width = 35
        ws_honorarios.column_dimensions['B'].width = 10
        ws_honorarios.column_dimensions['C'].width = 15
        ws_honorarios.column_dimensions['D'].width = 12
        ws_honorarios.column_dimensions['E'].width = 30
        
        # Formatar coluna de valor como moeda
        for row in range(2, len(honorarios) + 2):
            ws_honorarios.cell(row=row, column=3).number_format = 'R$ #,##0.00'
        
        # ===== ABA 3: POR CLIENTE =====
        ws_clientes = wb.create_sheet("Por Cliente")
        
        headers_cli = ["Cliente", "Total", "Pagos", "Pendentes", "Valor Total", "Valor Pago", "Valor Pendente"]
        for col, header in enumerate(headers_cli, start=1):
            cell = ws_clientes.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        clientes = listar_clientes()
        row = 2
        
        for cliente in clientes:
            h_cliente = [h for h in honorarios if h['cliente_id'] == cliente['id']]
            if not h_cliente:
                continue
            
            total = len(h_cliente)
            pagos = sum(1 for h in h_cliente if h['status'] == 'PAGO')
            pendentes = total - pagos
            valor_total = sum(h['valor'] for h in h_cliente)
            valor_pago = sum(h['valor'] for h in h_cliente if h['status'] == 'PAGO')
            valor_pendente = valor_total - valor_pago
            
            ws_clientes.cell(row=row, column=1, value=cliente['nome'])
            ws_clientes.cell(row=row, column=2, value=total)
            ws_clientes.cell(row=row, column=3, value=pagos)
            ws_clientes.cell(row=row, column=4, value=pendentes)
            ws_clientes.cell(row=row, column=5, value=valor_total)
            ws_clientes.cell(row=row, column=6, value=valor_pago)
            ws_clientes.cell(row=row, column=7, value=valor_pendente)
            
            # Formato moeda
            ws_clientes.cell(row=row, column=5).number_format = 'R$ #,##0.00'
            ws_clientes.cell(row=row, column=6).number_format = 'R$ #,##0.00'
            ws_clientes.cell(row=row, column=7).number_format = 'R$ #,##0.00'
            
            row += 1
        
        # Ajustar larguras
        for col in range(1, 8):
            ws_clientes.column_dimensions[get_column_letter(col)].width = 18
        ws_clientes.column_dimensions['A'].width = 35
        
        # Salvar
        if not pasta_destino:
            pasta_destino = os.path.join(os.path.expanduser("~"), "Desktop")
        
        filepath = os.path.join(pasta_destino, f"relatorio_honorarios_{ano}.xlsx")
        wb.save(filepath)
        
        return True, filepath
        
    except ImportError:
        # Fallback para CSV se openpyxl não estiver instalado
        return False, "openpyxl não instalado. Use: pip install openpyxl"
    except Exception as e:
        return False, str(e)


def exportar_excel_cliente(dados_relatorio, pasta_destino=None):
    """
    Exporta relatório individual de um cliente para Excel formatado
    
    Args:
        dados_relatorio: Dict retornado por get_relatorio_cliente()
        pasta_destino: Pasta onde salvar o arquivo
    
    Returns:
        Tuple (sucesso: bool, caminho_ou_erro: str)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        cliente = dados_relatorio['cliente']
        honorarios = dados_relatorio['honorarios']
        resumo_anos = dados_relatorio['resumo_anos']
        total_geral = dados_relatorio['total_geral']
        total_pago = dados_relatorio['total_pago']
        saldo_devedor = dados_relatorio['saldo_devedor']
        
        wb = Workbook()
        
        # Estilos
        titulo_font = Font(bold=True, size=14, color="FFFFFF")
        titulo_fill = PatternFill(start_color="1A2957", end_color="1A2957", fill_type="solid")
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="C4A962", end_color="C4A962", fill_type="solid")
        pago_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        pendente_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
        devedor_fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ===== ABA 1: RESUMO DO CLIENTE =====
        ws_resumo = wb.active
        ws_resumo.title = "Resumo"
        
        # Cabeçalho
        codigo = cliente.get('codigo_interno') or str(cliente.get('id', ''))
        ws_resumo.merge_cells('A1:D1')
        ws_resumo['A1'] = f"Relatório Individual - {cliente['nome']}"
        ws_resumo['A1'].font = titulo_font
        ws_resumo['A1'].fill = titulo_fill
        ws_resumo['A1'].alignment = Alignment(horizontal='center')
        
        # Dados do cliente
        ws_resumo['A3'] = "Código:"
        ws_resumo['A3'].font = Font(bold=True)
        ws_resumo['B3'] = codigo
        
        ws_resumo['A4'] = "Nome:"
        ws_resumo['A4'].font = Font(bold=True)
        ws_resumo['B4'] = cliente['nome']
        
        if cliente.get('cnpj'):
            ws_resumo['A5'] = "CNPJ/CPF:"
            ws_resumo['A5'].font = Font(bold=True)
            ws_resumo['B5'] = cliente['cnpj']
        
        # Totais
        ws_resumo['A7'] = "TOTAIS GERAIS"
        ws_resumo['A7'].font = Font(bold=True, size=12)
        
        metricas = [
            ("Total Geral", total_geral, None),
            ("Total Pago", total_pago, pago_fill),
            ("Saldo Devedor", saldo_devedor, devedor_fill),
        ]
        
        for i, (label, valor, fill) in enumerate(metricas, start=8):
            ws_resumo[f'A{i}'] = label
            ws_resumo[f'A{i}'].font = Font(bold=True)
            ws_resumo[f'B{i}'] = valor
            ws_resumo[f'B{i}'].number_format = 'R$ #,##0.00'
            if fill:
                ws_resumo[f'B{i}'].fill = fill
                ws_resumo[f'B{i}'].font = Font(bold=True, color="FFFFFF")
        
        # Ajustar larguras
        ws_resumo.column_dimensions['A'].width = 20
        ws_resumo.column_dimensions['B'].width = 25
        
        # Data de geração
        ws_resumo['A12'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws_resumo['A12'].font = Font(italic=True, size=9)
        
        # ===== ABA 2: HONORÁRIOS =====
        ws_honorarios = wb.create_sheet("Honorários")
        
        MESES = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez", "13º"]
        
        headers = ["Ano", "Mês", "Valor", "Status", "Observação"]
        for col, header in enumerate(headers, start=1):
            cell = ws_honorarios.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        for i, h in enumerate(honorarios, start=2):
            mes_nome = MESES[h['mes']-1] if h['mes'] <= 13 else str(h['mes'])
            
            ws_honorarios.cell(row=i, column=1, value=h['ano'])
            ws_honorarios.cell(row=i, column=2, value=mes_nome)
            ws_honorarios.cell(row=i, column=3, value=h['valor'])
            ws_honorarios.cell(row=i, column=4, value=h['status'])
            ws_honorarios.cell(row=i, column=5, value=h.get('observacao') or '')
            
            # Formatar valor
            ws_honorarios.cell(row=i, column=3).number_format = 'R$ #,##0.00'
            
            # Cor do status
            if h['status'] == 'PAGO':
                ws_honorarios.cell(row=i, column=4).fill = pago_fill
                ws_honorarios.cell(row=i, column=4).font = Font(color="FFFFFF")
            elif h['status'] == 'PENDENTE':
                ws_honorarios.cell(row=i, column=4).fill = pendente_fill
                ws_honorarios.cell(row=i, column=4).font = Font(color="FFFFFF")
            
            # Bordas
            for col in range(1, 6):
                ws_honorarios.cell(row=i, column=col).border = thin_border
        
        # Ajustar larguras
        ws_honorarios.column_dimensions['A'].width = 10
        ws_honorarios.column_dimensions['B'].width = 10
        ws_honorarios.column_dimensions['C'].width = 15
        ws_honorarios.column_dimensions['D'].width = 12
        ws_honorarios.column_dimensions['E'].width = 40
        
        # ===== ABA 3: RESUMO POR ANO =====
        ws_anos = wb.create_sheet("Por Ano")
        
        headers_ano = ["Ano", "Total", "Pagos", "Pendentes", "Valor Total", "Valor Pago", "Valor Pendente"]
        for col, header in enumerate(headers_ano, start=1):
            cell = ws_anos.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        for i, r in enumerate(resumo_anos, start=2):
            ws_anos.cell(row=i, column=1, value=r['ano'])
            ws_anos.cell(row=i, column=2, value=r['total'])
            ws_anos.cell(row=i, column=3, value=r['pagos'])
            ws_anos.cell(row=i, column=4, value=r['pendentes'])
            ws_anos.cell(row=i, column=5, value=r['valor_total'] or 0)
            ws_anos.cell(row=i, column=6, value=r['valor_pago'] or 0)
            ws_anos.cell(row=i, column=7, value=r['valor_pendente'] or 0)
            
            # Formatar moedas
            ws_anos.cell(row=i, column=5).number_format = 'R$ #,##0.00'
            ws_anos.cell(row=i, column=6).number_format = 'R$ #,##0.00'
            ws_anos.cell(row=i, column=7).number_format = 'R$ #,##0.00'
            
            # Bordas
            for col in range(1, 8):
                ws_anos.cell(row=i, column=col).border = thin_border
        
        # Ajustar larguras
        for col in range(1, 8):
            ws_anos.column_dimensions[get_column_letter(col)].width = 15
        ws_anos.column_dimensions['A'].width = 10
        
        # Salvar
        if not pasta_destino:
            pasta_destino = os.path.join(os.path.expanduser("~"), "Desktop")
        
        nome_limpo = "".join(c for c in cliente['nome'] if c.isalnum() or c in " -_").strip()[:30]
        filepath = os.path.join(pasta_destino, f"relatorio_{codigo}_{nome_limpo}.xlsx")
        wb.save(filepath)
        
        return True, filepath
        
    except ImportError:
        return False, "openpyxl não instalado. Use: pip install openpyxl"
    except Exception as e:
        return False, str(e)
