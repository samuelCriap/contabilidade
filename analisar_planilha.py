"""
Debug: verificar quantas linhas têm código numérico
"""
import openpyxl
print("Carregando...")
wb = openpyxl.load_workbook('2025.xlsx', data_only=True)
ws = wb['2025']

# Contar linhas com código
codigos = []
for r in range(1, ws.max_row + 1):
    codigo = ws.cell(row=r, column=1).value
    if codigo:
        try:
            int(codigo)
            nome = ws.cell(row=r, column=2).value
            codigos.append((r, codigo, nome))
        except:
            pass

print(f"Total de linhas com código numérico: {len(codigos)}")
print("\nPrimeiros 10:")
for r, c, n in codigos[:10]:
    print(f"  Linha {r}: {c} - {n}")
print("\nÚltimos 10:")
for r, c, n in codigos[-10:]:
    print(f"  Linha {r}: {c} - {n}")
