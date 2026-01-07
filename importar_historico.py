"""
Script de Importação de Histórico de Pagamentos
Importa dados de planilhas Excel (2025.xlsx, etc.) para o banco de dados.

Estrutura esperada da planilha:
- Coluna A: Código do cliente
- Coluna B: Nome do cliente  
- Colunas C-O: Meses (Jan, Fev, Mar, Abr, Mai, Jun, Jul, Ago, Set, Out, Nov, 13º, Dez)
- Cada cliente ocupa 4 linhas:
  - Linha 1: Nome (mesclado, ignorar)
  - Linha 2: Valor (ex: R$645,00)
  - Linha 3: Data do pagamento (ex: 09/dez)
  - Linha 4: Forma de pagamento (ex: DINHEIRO)
"""

import openpyxl
import os
import sys
import re
from datetime import datetime

# Adicionar path do projeto (honorarios folder)
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), 'honorarios'))

from database import get_connection, listar_clientes

# Mapeamento de colunas para meses (C=3 até O=15)
# C=Jan(1), D=Fev(2), E=Mar(3), F=Abr(4), G=Mai(5), H=Jun(6), 
# I=Jul(7), J=Ago(8), K=Set(9), L=Out(10), M=Nov(11), N=13º(13), O=Dez(12)
COLUNAS_MESES = {
    3: 1,   # C = Janeiro
    4: 2,   # D = Fevereiro
    5: 3,   # E = Março
    6: 4,   # F = Abril
    7: 5,   # G = Maio
    8: 6,   # H = Junho
    9: 7,   # I = Julho
    10: 8,  # J = Agosto
    11: 9,  # K = Setembro
    12: 10, # L = Outubro
    13: 11, # M = Novembro
    14: 13, # N = 13º Salário
    15: 12, # O = Dezembro
}

MESES_ABREV = {
    'jan': 1, 'janeiro': 1,
    'fev': 2, 'fevereiro': 2,
    'mar': 3, 'março': 3, 'marco': 3,
    'abr': 4, 'abril': 4,
    'mai': 5, 'maio': 5,
    'jun': 6, 'junho': 6,
    'jul': 7, 'julho': 7,
    'ago': 8, 'agosto': 8,
    'set': 9, 'setembro': 9,
    'out': 10, 'outubro': 10,
    'nov': 11, 'novembro': 11,
    'dez': 12, 'dezembro': 12,
}


def parse_valor(valor_str):
    """Converte string de valor para float"""
    if not valor_str:
        return None
    try:
        if isinstance(valor_str, (int, float)):
            return float(valor_str)
        # Remove R$, espaços e converte vírgula para ponto
        valor = str(valor_str).replace('R$', '').replace(' ', '').replace('.', '').replace(',', '.')
        return float(valor)
    except:
        return None


def parse_data(data_str, ano):
    """Converte string de data para formato YYYY-MM-DD"""
    if not data_str:
        return None
    try:
        data_str = str(data_str).lower().strip()
        
        # Formato: DD/mes (ex: 09/dez)
        match = re.match(r'(\d{1,2})[/\-](\w+)', data_str)
        if match:
            dia = int(match.group(1))
            mes_str = match.group(2)
            mes = MESES_ABREV.get(mes_str[:3], None)
            if mes:
                return f"{ano}-{mes:02d}-{dia:02d}"
        
        # Formato: DD/MM/YYYY ou DD/MM
        match = re.match(r'(\d{1,2})[/\-](\d{1,2})(?:[/\-](\d{2,4}))?', data_str)
        if match:
            dia = int(match.group(1))
            mes = int(match.group(2))
            ano_data = int(match.group(3)) if match.group(3) else ano
            if ano_data < 100:
                ano_data += 2000
            return f"{ano_data}-{mes:02d}-{dia:02d}"
        
        return None
    except:
        return None


def importar_planilha(caminho_arquivo, ano):
    """Importa dados de uma planilha para o banco de dados"""
    print(f"\n{'='*50}")
    print(f"Importando: {caminho_arquivo} (Ano: {ano})")
    print('='*50)
    
    # Carregar planilha
    wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
    ws = wb.active
    
    # Carregar clientes do banco
    clientes_list = listar_clientes()
    clientes_db = {}
    clientes_nome = {}
    for c in clientes_list:
        codigo = c['codigo_interno'] if c['codigo_interno'] else None
        if codigo:
            clientes_db[str(codigo)] = c['id']
        clientes_nome[c['nome'].upper()] = c['id']
    
    conn = get_connection()
    cursor = conn.cursor()
    
    importados = 0
    erros = 0
    linha = 2  # Começar na linha 2 (pular cabeçalho)
    
    # Processar em blocos de 4 linhas (1 cliente por bloco)
    while linha <= ws.max_row:
        # Linha 1 do bloco: Código e Nome
        codigo_cliente = ws.cell(row=linha, column=1).value
        nome_cliente = ws.cell(row=linha, column=2).value
        
        # Pular linhas vazias ou que não são clientes
        if not codigo_cliente or not nome_cliente:
            linha += 4
            continue
            
        # Tentar converter código para inteiro (ignorar textos como "Data", "Pagamento")
        try:
            codigo_int = int(codigo_cliente)
        except (ValueError, TypeError):
            linha += 4
            continue
        
        # Encontrar cliente no banco
        cliente_id = None
        codigo_str = str(codigo_cliente).strip() if codigo_cliente else None
        nome_str = str(nome_cliente).strip() if nome_cliente else "Cliente " + codigo_str
        
        if codigo_str and codigo_str in clientes_db:
            cliente_id = clientes_db[codigo_str]
        elif nome_str:
            nome_upper = nome_str.upper()
            if nome_upper in clientes_nome:
                cliente_id = clientes_nome[nome_upper]
        
        if not cliente_id:
            # Criar cliente automaticamente (ex-cliente, inativo)
            try:
                cursor.execute("""
                    INSERT INTO clientes (nome, codigo_interno, email, ativo)
                    VALUES (?, ?, '', 0)
                """, (nome_str, codigo_str))
                cliente_id = cursor.lastrowid
                clientes_db[codigo_str] = cliente_id
                clientes_nome[nome_str.upper()] = cliente_id
                print(f"  [NOVO] Cliente criado: {codigo_str} - {nome_str}")
            except Exception as e:
                print(f"  [ERRO] Falha ao criar cliente {codigo_str} - {nome_str}: {e}")
                linha += 4
                erros += 1
                continue
        
        # Estrutura correta:
        # Linha N: Código | Nome | VALORES
        # Linha N+1: (mesclado)
        # Linha N+2: _ | "Data" | Datas
        # Linha N+3: _ | "Pagamento" | Formas
        
        for col, mes in COLUNAS_MESES.items():
            # Valor (mesma linha do código/nome)
            valor_cell = ws.cell(row=linha, column=col).value
            valor = parse_valor(valor_cell)
            
            if not valor or valor <= 0:
                continue
            
            # Data (linha + 2)
            data_cell = ws.cell(row=linha + 2, column=col).value
            data_pagamento = parse_data(data_cell, ano)
            
            # Forma de pagamento (linha + 3)
            forma_cell = ws.cell(row=linha + 3, column=col).value
            forma_pagamento = str(forma_cell).strip() if forma_cell else None
            
            # Determinar status
            status = 'PAGO' if data_pagamento or forma_pagamento else 'PENDENTE'
            
            # Inserir no banco
            try:
                cursor.execute("""
                    INSERT INTO honorarios (cliente_id, ano, mes, valor, status, data_pagamento, forma_pagamento)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (cliente_id, ano, mes, valor, status, data_pagamento, forma_pagamento))
                importados += 1
            except Exception as e:
                print(f"  [ERRO] Cliente {codigo_cliente}, Mês {mes}: {e}")
                erros += 1
        
        linha += 4
    
    conn.commit()
    conn.close()
    
    print(f"\nResultado: {importados} registros importados, {erros} erros")
    return importados, erros


def limpar_honorarios():
    """Remove todos os honorários existentes"""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM honorarios")
    count = cursor.rowcount
    conn.commit()
    conn.close()
    print(f"Removidos {count} registros de honorários existentes.")
    return count


def main():
    print("\n" + "="*60)
    print("   IMPORTAÇÃO DE HISTÓRICO DE PAGAMENTOS")
    print("="*60)
    
    # 1. Limpar honorários existentes
    print("\nApagando honorários existentes...")
    limpar_honorarios()
    
    # 2. Importar planilhas
    arquivos = [
        ('2025.xlsx', 2025),
        # Adicione mais arquivos conforme necessário:
        # ('2024.xlsx', 2024),
        # ('2023.xlsx', 2023),
    ]
    
    total_importados = 0
    total_erros = 0
    
    for arquivo, ano in arquivos:
        if os.path.exists(arquivo):
            imp, err = importar_planilha(arquivo, ano)
            total_importados += imp
            total_erros += err
        else:
            print(f"\n[AVISO] Arquivo não encontrado: {arquivo}")
    
    print("\n" + "="*60)
    print(f"   IMPORTAÇÃO CONCLUÍDA")
    print(f"   Total importados: {total_importados}")
    print(f"   Total erros: {total_erros}")
    print("="*60)


if __name__ == "__main__":
    main()
