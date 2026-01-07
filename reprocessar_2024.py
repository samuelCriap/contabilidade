"""
Script para processar 2024 diretamente da planilha (aba 2024 de 2025.xlsx)
"""
import openpyxl
import os
import sys
import re
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), 'honorarios'))
from database import get_connection

COLUNAS_MESES = {
    3: 1, 4: 2, 5: 3, 6: 4, 7: 5, 8: 6,
    9: 7, 10: 8, 11: 9, 12: 10, 13: 11, 14: 13, 15: 12
}

MESES_ABREV = {
    'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4, 'mai': 5, 'jun': 6,
    'jul': 7, 'ago': 8, 'set': 9, 'out': 10, 'nov': 11, 'dez': 12,
}

def parse_data(data_str, ano):
    if not data_str:
        return None
    try:
        data_str = str(data_str).lower().strip()
        match = re.match(r'(\d{1,2})[/\\-](\w+)', data_str)
        if match:
            dia = int(match.group(1))
            mes_str = match.group(2)
            mes = MESES_ABREV.get(mes_str[:3], None)
            if mes:
                return f"{ano}-{mes:02d}-{dia:02d}"
        match = re.match(r'(\d{1,2})[/\\-](\d{1,2})(?:[/\\-](\d{2,4}))?', data_str)
        if match:
            dia = int(match.group(1))
            mes = int(match.group(2))
            ano_data = int(match.group(3)) if match.group(3) else ano
            if ano_data < 100:
                ano_data += 2000
            return f"{ano_data}-{mes:02d}-{dia:02d}"
        return None
    except:
        return None

def parse_valor(valor_cell):
    if not valor_cell:
        return None
    if isinstance(valor_cell, (int, float)):
        return float(valor_cell)
    valor_str = str(valor_cell).replace('R$', '').replace(' ', '').replace('.', '').replace(',', '.')
    try:
        return float(valor_str)
    except:
        return None

def processar_ano(arquivo, aba, ano):
    print(f"\n📂 Processando: {arquivo} - Aba: {aba} - Ano: {ano}")
    
    conn = get_connection()
    cursor = conn.cursor()
    
    # 1. Apagar todos os honorários do ano
    print(f"\n🗑️  Apagando todos os honorários de {ano}...")
    cursor.execute("DELETE FROM honorarios WHERE ano = ?", (ano,))
    deletados = cursor.rowcount
    conn.commit()
    print(f"  ✅ {deletados} registros deletados")
    
    # 2. Carregar clientes do banco
    cursor.execute("SELECT id, codigo_interno, UPPER(nome) as nome FROM clientes")
    clientes_db = {}
    clientes_nome = {}
    for c in cursor.fetchall():
        if c['codigo_interno']:
            clientes_db[str(c['codigo_interno']).strip()] = c['id']
        clientes_nome[c['nome']] = c['id']
    
    # 3. Carregar planilha
    print("\n📂 Carregando planilha...")
    wb = openpyxl.load_workbook(arquivo, data_only=True)
    ws = wb[aba]
    print(f"  ✅ Planilha carregada ({ws.max_row} linhas)")
    
    # 4. Encontrar linhas com código numérico
    linhas_clientes = []
    for r in range(1, ws.max_row + 1):
        codigo = ws.cell(row=r, column=1).value
        if codigo:
            try:
                int(codigo)
                linhas_clientes.append(r)
            except:
                pass
    print(f"  ✅ Encontradas {len(linhas_clientes)} linhas com clientes")
    
    # 5. Processar cada cliente
    print("\n🔄 Processando clientes e honorários...")
    clientes_processados = 0
    pagos = 0
    pendentes = 0
    novos_clientes = 0
    
    for linha in linhas_clientes:
        codigo_cliente = ws.cell(row=linha, column=1).value
        nome_cliente = ws.cell(row=linha, column=2).value
        
        if not nome_cliente:
            continue
        
        codigo_str = str(codigo_cliente).strip()
        nome_str = str(nome_cliente).strip()
        
        # Encontrar ou criar cliente
        cliente_id = clientes_db.get(codigo_str) or clientes_nome.get(nome_str.upper())
        
        if not cliente_id:
            cursor.execute("""
                INSERT INTO clientes (nome, codigo_interno, email, ativo)
                VALUES (?, ?, '', 0)
            """, (nome_str, codigo_str))
            cliente_id = cursor.lastrowid
            clientes_db[codigo_str] = cliente_id
            clientes_nome[nome_str.upper()] = cliente_id
            novos_clientes += 1
        
        # Encontrar valor padrão do cliente
        valor_padrao = None
        for col in COLUNAS_MESES.keys():
            v = parse_valor(ws.cell(row=linha, column=col).value)
            if v and v > 0:
                valor_padrao = v
                break
        
        # Processar cada mês
        for col, mes in COLUNAS_MESES.items():
            valor = parse_valor(ws.cell(row=linha, column=col).value)
            data_cell = ws.cell(row=linha + 2, column=col).value
            forma_cell = ws.cell(row=linha + 3, column=col).value
            
            data_pagamento = parse_data(data_cell, ano)
            forma_pagamento = str(forma_cell).strip() if forma_cell else None
            
            if valor and valor > 0:
                # Tem valor = PAGO
                cursor.execute("""
                    INSERT INTO honorarios (cliente_id, ano, mes, valor, status, data_pagamento, forma_pagamento)
                    VALUES (?, ?, ?, ?, 'PAGO', ?, ?)
                """, (cliente_id, ano, mes, valor, data_pagamento, forma_pagamento))
                pagos += 1
            else:
                # Sem valor = PENDENTE com valor padrão
                if valor_padrao:
                    cursor.execute("""
                        INSERT INTO honorarios (cliente_id, ano, mes, valor, status)
                        VALUES (?, ?, ?, ?, 'PENDENTE')
                    """, (cliente_id, ano, mes, valor_padrao))
                    pendentes += 1
        
        clientes_processados += 1
    
    conn.commit()
    conn.close()
    
    print(f"\n  ✅ Clientes processados: {clientes_processados}")
    print(f"  ✅ Novos clientes criados: {novos_clientes}")
    print(f"  ✅ Honorários PAGOS: {pagos}")
    print(f"  ✅ Honorários PENDENTES: {pendentes}")
    
    return clientes_processados, novos_clientes, pagos, pendentes

if __name__ == "__main__":
    print("\n" + "="*60)
    print("   PROCESSAMENTO DIRETO DA PLANILHA 2024")
    print("="*60)
    
    processar_ano("2025.xlsx", "2024", 2024)
    
    print("\n" + "="*60)
    print("   CONCLUÍDO!")
    print("="*60)
